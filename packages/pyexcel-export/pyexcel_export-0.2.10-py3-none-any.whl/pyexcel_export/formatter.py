import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.copier import WorksheetCopy

import logging
import json
from io import BytesIO
from collections import OrderedDict
from pathlib import Path

from .serialize import MyEncoder
from .defaults import Meta

debug_logger = logging.getLogger('debug')


class ExcelFormatter:
    def __init__(self, template_file=None):
        """

        :param str|Path|None template_file:
        """
        is_read = False

        if template_file:
            if not isinstance(template_file, Path):
                template_file = Path(template_file)

            if template_file.exists():
                self.styled_wb = openpyxl.load_workbook(template_file)
                self.to_stylesheets(self.styled_wb)

                is_read = True

        if not is_read:
            self.styled_wb = openpyxl.Workbook()
            self.styled_wb.active.title = '_template'

    @property
    def data(self):
        output = BytesIO()
        self.styled_wb.save(output)

        return output

    @data.setter
    def data(self, _styles):
        self.styled_wb = openpyxl.load_workbook(_styles)

    def save(self, raw_data, out_file, meta=None, retain_meta=True):
        """

        :param raw_data:
        :param str|Path out_file:
        :param meta:
        :param retain_meta:
        :return:
        """
        if not isinstance(out_file, Path):
            out_file = Path(out_file)

        if not meta:
            meta = Meta()

        if '_styles' in meta.keys():
            self.data = meta['_styles']
            meta['_styles'] = self.data

        if out_file.exists():
            wb = openpyxl.load_workbook(out_file)
            self.to_stylesheets(wb)
            self.append_styled_sheets(wb)

            original_sheet_names = []
            extraneous_sheet_names = []
        else:
            wb = self.styled_wb
            original_sheet_names = wb.sheetnames
            extraneous_sheet_names = []

        extraneous_sheet_names.append('_template')
        inserted_sheets = []

        if not retain_meta:
            extraneous_sheet_names.append('_meta')
        else:
            if '_meta' in original_sheet_names:
                wb.remove(wb['_meta'])

            self.create_styled_sheet(wb, '_meta', 0)

            meta_matrix = []

            for k, v in meta.excel_matrix:
                if not k.startswith('_'):
                    if isinstance(v, (dict, OrderedDict)):
                        v = json.dumps(v, cls=MyEncoder)
                    meta_matrix.append([k, v])

            self.fill_matrix(wb['_meta'], meta_matrix, rules=meta)

        if '_meta' in raw_data.keys():
            raw_data.pop('_meta')

        for sheet_name, cell_matrix in raw_data.items():
            if sheet_name not in original_sheet_names:
                self.create_styled_sheet(wb, sheet_name)
                inserted_sheets.append(sheet_name)
            else:
                if meta.get('allow_table_hiding', True) in (True, 'true'):
                    if not sheet_name.startswith('_'):
                        self.fill_matrix(wb[sheet_name], cell_matrix, rules=meta)
                    else:
                        for i, cell in enumerate(next(wb['_meta'].iter_cols())):
                            if not cell.value:
                                matrix = [[sheet_name]]
                                matrix.extend(cell_matrix)

                                self.fill_matrix(wb['_meta'], matrix, start_row=i+1, rules=meta)
                                break
                else:
                    self.fill_matrix(wb[sheet_name], cell_matrix, rules=meta)

            ws = wb[sheet_name]
            for row_num, row in enumerate(cell_matrix):
                for col_num, value in enumerate(row):
                    if isinstance(value, (dict, OrderedDict)):
                        value = json.dumps(value, cls=MyEncoder)

                    ws.cell(column=(col_num + 1),
                            row=(row_num + 1),
                            value=value)

        for sheet_name in extraneous_sheet_names:
            if sheet_name in wb.sheetnames:
                wb.remove(wb[sheet_name])

        for sheet_name in wb.sheetnames:
            if (sheet_name.startswith('_') and sheet_name != '_meta') or self.is_empty_sheet(wb[sheet_name]):
                wb.remove(wb[sheet_name])

        wb.save(out_file)

    @staticmethod
    def create_styled_sheet(wb, sheet_name, pos: int=None):
        wb.create_sheet(sheet_name, pos)
        if '_template' in wb.sheetnames and sheet_name != '_template':
            WorksheetCopy(wb['_template'], wb[sheet_name]).copy_worksheet()
            # wb[sheet_name].copy_worksheet(wb['_template'])

    def append_styled_sheets(self, wb):
        if '_template' not in wb.sheetnames:
            wb.create_sheet('_template')
            if '_template' in self.styled_wb.sheetnames:
                WorksheetCopy(self.styled_wb['_template'], wb['_template']).copy_worksheet()
                # wb['_template'].copy_worksheet(wb['_template'])

        return wb

    @staticmethod
    def to_stylesheets(wb):
        for ws in wb:
            for row in ws:
                for cell in row:
                    cell.value = None

        return wb

    @staticmethod
    def fill_matrix(ws, cell_matrix, start_row=0, rules=None):
        for row_num, row in enumerate(cell_matrix):
            for col_num, value in enumerate(row):
                if isinstance(value, (dict, OrderedDict)):
                    value = json.dumps(value, cls=MyEncoder)

                ws.cell(column=(col_num + 1),
                        row=(row_num + start_row + 1),
                        value=value)

        if rules is not None:
            if rules.get('has_header', False) in (True, 'true') \
                    and rules.get('freeze_header', False) in (True, 'true'):
                if ws.title != '_meta':
                    ws.freeze_panes = 'A2'
            if rules.get('col_width_fit_param_keys', False) in (True, 'true'):
                width = max([len(str(cell.value)) for cell in next(ws.iter_cols())])
                ws.column_dimensions['A'].width = width + 2
            if rules.get('col_width_fit_ids', False) in (True, 'true'):
                for i, header_cell in enumerate(next(ws.iter_rows())):
                    header_item = header_cell.value
                    if header_item and header_item.endswith('id'):
                        col_letter = get_column_letter(i + 1)
                        width = max([len(str(cell.value)) for cell in list(ws.iter_cols())[i]])
                        ws.column_dimensions[col_letter].width = width + 2

    @staticmethod
    def is_empty_sheet(ws):
        def is_not_empty():
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value:
                        return True

            return False

        return not is_not_empty()
